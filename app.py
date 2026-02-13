import os
import re
import io
import unicodedata
from functools import wraps
import pandas as pd
from datetime import datetime
import warnings
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.sql import text
from sqlalchemy import or_
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Tắt cảnh báo UserWarning của openpyxl (thường gặp khi đọc file có Data Validation)
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# 1. Tải biến môi trường từ file .env
load_dotenv()

# 2. Khởi tạo ứng dụng Flask
app = Flask(__name__)

# 3. Cấu hình Database
# Lấy chuỗi kết nối từ .env. Nếu không có sẽ báo lỗi.
db_url = os.getenv("DATABASE_URL")
if not db_url:
    raise ValueError("Vui lòng thiết lập DATABASE_URL trong file .env")

# Cấu hình SQLAlchemy
app.config['SECRET_KEY'] = 'a_dev_secret_key_that_should_be_changed_in_production'
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_recycle': 280, # Giúp duy trì kết nối với MySQL (tránh lỗi timeout)
    'pool_pre_ping': True # Kiểm tra kết nối trước khi gửi lệnh
}

# Khởi tạo đối tượng DB
db = SQLAlchemy(app)

# Cấu hình Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login' # Tên hàm view cho trang login

# Định nghĩa Model cho bảng employees_tbs
class Employee(db.Model):
    __tablename__ = 'employees_tbs'
    id = db.Column(db.Integer, primary_key=True)
    employee_code = db.Column(db.String(20), unique=True, nullable=False)
    full_name = db.Column(db.String(100), nullable=False)
    position = db.Column(db.String(50))
    employee_type = db.Column(db.String(30))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime)
    masl=db.Column(db.String(50))
    info=db.Column(db.String(200))

class Customer(db.Model):
    __tablename__ = 'customers'
    id = db.Column(db.Integer, primary_key=True)
    customer_code = db.Column(db.String(20), unique=True, nullable=False)
    customer_name = db.Column(db.String(100), nullable=False)

class CustomerAccount(db.Model):
    __tablename__ = 'customer_accounts'
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customers.id'), nullable=False)
    account_code = db.Column(db.String(30), nullable=False)
    account_name = db.Column(db.String(100))
    is_active = db.Column(db.Boolean, default=True)
    customer = db.relationship('Customer', backref=db.backref('accounts', lazy=True))

class AccountTask(db.Model):
    __tablename__ = 'account_tasks'
    id = db.Column(db.Integer, primary_key=True)
    account_id = db.Column(db.Integer, db.ForeignKey('customer_accounts.id'), nullable=False)
    task_code = db.Column(db.String(20), nullable=False)
    task_name = db.Column(db.String(50), nullable=False)
    account = db.relationship('CustomerAccount', backref=db.backref('tasks', lazy=True))

class AccountConversionIndex(db.Model):
    __tablename__ = 'account_conversion_index'
    id = db.Column(db.Integer, primary_key=True)
    account_id = db.Column(db.Integer, db.ForeignKey('customer_accounts.id'), nullable=False)
    task_id = db.Column(db.Integer, db.ForeignKey('account_tasks.id'), nullable=False)
    conversion_index = db.Column(db.Numeric(8, 3), nullable=False)
    unit = db.Column(db.String(20))
    effective_from = db.Column(db.Date, nullable=False)
    effective_to = db.Column(db.Date)
    account = db.relationship('CustomerAccount', backref=db.backref('conversion_indices', lazy=True))
    task = db.relationship('AccountTask', backref=db.backref('conversion_indices', lazy=True))

class LaborProductivity(db.Model):
    __tablename__ = 'labor_productivity'
    id = db.Column(db.Integer, primary_key=True)
    work_date = db.Column(db.Date)
    ref_no = db.Column(db.String(50))
    productivity_value = db.Column(db.Float)
    tally_id = db.Column(db.String(100))
    xenang_id = db.Column(db.String(100))
    congnhan1_id = db.Column(db.String(100))
    congnhan2_id = db.Column(db.String(100))
    congnhan3_id = db.Column(db.String(100))
    congnhan4_id = db.Column(db.String(100))
    congnhan5_id = db.Column(db.String(100))
    congnhan6_id = db.Column(db.String(100))
    task_id = db.Column(db.String(100))
    account_id = db.Column(db.String(100))
    customer_id = db.Column(db.String(100))
    unit= db.Column(db.String(50))
    conversion_index= db.Column(db.Float)
    quantity= db.Column(db.Float)

class LaborProductivityTemp(db.Model):
    __tablename__ = 'labor_productivity_temp'
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date)
    container_no = db.Column(db.String(50))
    cbm = db.Column(db.Float)
    tally = db.Column(db.String(100))
    lift_truck = db.Column(db.String(100))
    worker_1 = db.Column(db.String(100), nullable=True)
    worker_2 = db.Column(db.String(100), nullable=True)
    worker_3 = db.Column(db.String(100), nullable=True)
    worker_4 = db.Column(db.String(100), nullable=True)
    worker_5 = db.Column(db.String(100), nullable=True)
    worker_6 = db.Column(db.String(100), nullable=True)
    task = db.Column(db.String(100))
    account = db.Column(db.String(100))
    customer = db.Column(db.String(100))

class LaborProductivityStaff(db.Model):
    __tablename__ = 'labor_productivity_staff'
    id = db.Column(db.Integer, primary_key=True)
    productivity_id = db.Column(db.Integer, nullable=False) # FK tới labor_productivity.id
    employee_id = db.Column(db.Integer, nullable=False)
    role = db.Column(db.String(30), nullable=False)
    ratio = db.Column(db.Float, default=1.0) # Tương ứng với DEFAULT 1.0 trong SQL

class SystemSetting(db.Model):
    __tablename__ = 'system_settings'
    id = db.Column(db.Integer, primary_key=True)
    key_name = db.Column(db.String(50), unique=True, nullable=False)
    value = db.Column(db.String(255))

class User(UserMixin, db.Model):
    __tablename__ = 'users_tbs'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    full_name = db.Column(db.String(100))
    role = db.Column(db.Enum('ADMIN', 'UPDATE', 'VIEW'), nullable=False, default='VIEW')
    is_active = db.Column(db.Boolean, default=True)
    can_export = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

# Định nghĩa Model cho bảng chucvu_tbs
class ChucVu(db.Model):
    __tablename__ = 'chucvu_tbs'
    id = db.Column(db.Integer, primary_key=True)
    ten_chuc_vu = db.Column(db.String(100), nullable=False, unique=True)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# --- Decorators phân quyền ---
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.role != 'ADMIN':
            flash('Bạn không có quyền truy cập trang này.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def update_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.role not in ['ADMIN', 'UPDATE']:
            flash('Bạn không có quyền truy cập trang này.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def view_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.role not in ['ADMIN', 'VIEW']:
            flash('Bạn không có quyền truy cập trang này.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            if not user.is_active:
                flash('Tài khoản này đã bị khóa.', 'danger')
                return redirect(url_for('login'))
                
            login_user(user)
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        else:
            flash('Tên đăng nhập hoặc mật khẩu không đúng.', 'danger')
            
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    db_status = "Chưa kết nối"
    version = ""
    
    try:
        # 4. Thử kết nối và truy vấn đơn giản để kiểm tra
        # Sử dụng text() để thực thi câu lệnh SQL thô an toàn
        result = db.session.execute(text("SELECT VERSION()"))
        version = result.scalar()
        db_status = "Kết nối thành công tới Aiven MySQL!"
    except Exception as e:
        db_status = f"Lỗi kết nối: {str(e)}"
        print(e) # In lỗi ra terminal để debug

    # 5. Render file HTML và truyền dữ liệu qua
    return render_template('index.html', status=db_status, version=version)

@app.route('/nhan-vien', methods=['GET', 'POST'])
@login_required
@admin_required
def nhan_vien():
    if request.method == 'POST':
        code_to_check = request.form['employee_code']
        return_url = request.form.get('return_url')
        
        # Kiểm tra trùng lặp mã nhân viên khi thêm mới
        existing_employee = Employee.query.filter_by(employee_code=code_to_check).first()
        if existing_employee:
            flash(f'Mã nhân viên "{code_to_check}" đã tồn tại. Vui lòng sử dụng mã khác.', 'danger')
            if return_url: return redirect(return_url)
            return redirect(url_for('nhan_vien'))

        try:
            new_emp = Employee(
                employee_code=code_to_check,
                full_name=request.form['full_name'].title(),
                position=request.form.get('position'),
                employee_type=request.form.get('employee_type'),
                masl=request.form.get('masl'),
                info=request.form.get('info'),
                is_active=True if request.form.get('is_active') else False,
                created_at=datetime.now()
            )
            db.session.add(new_emp)
            db.session.commit()
            flash('Thêm nhân viên mới thành công!', 'success')
        except Exception as e:
            print(f"Lỗi thêm nhân viên: {e}")
            db.session.rollback()
            flash(f'Lỗi khi thêm nhân viên: {e}', 'danger')
        
        if return_url: return redirect(return_url)
        return redirect(url_for('nhan_vien'))

    page = request.args.get('page', 1, type=int)
    search_term = request.args.get('search_masl', '')
    per_page = 20
    employees = None    
    try:
        # Xây dựng câu truy vấn cơ bản
        query = Employee.query

        # Áp dụng bộ lọc tìm kiếm nếu có
        if search_term:
            query = query.filter(Employee.masl.ilike(f'%{search_term}%'))

        # Sắp xếp nhân viên mới nhất lên đầu và phân trang
        employees = query.order_by(Employee.id.desc()).paginate(page=page, per_page=per_page, error_out=False)
    except Exception as e:
        print(f"Lỗi truy vấn nhân viên: {e}")
    return render_template('nhanvien.html', employees=employees, search_term=search_term)

@app.route('/nhan-vien/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_nhan_vien(id):
    emp = Employee.query.get_or_404(id)
    new_code = request.form['employee_code']

    # Kiểm tra xem mã mới có bị trùng với nhân viên khác không
    existing_employee = Employee.query.filter(Employee.employee_code == new_code, Employee.id != id).first()
    if existing_employee:
        flash(f'Mã nhân viên "{new_code}" đã được sử dụng bởi một nhân viên khác.', 'danger')
        return redirect(url_for('nhan_vien'))
    try:
        emp.employee_code = new_code
        emp.full_name = request.form['full_name'].title()
        emp.position = request.form.get('position')
        emp.employee_type = request.form.get('employee_type')
        emp.masl = request.form.get('masl')
        emp.info = request.form.get('info')
        emp.is_active = True if request.form.get('is_active') else False
        db.session.commit()
        flash('Cập nhật thông tin nhân viên thành công!', 'success')
    except Exception as e:
        print(f"Lỗi sửa nhân viên: {e}")
        db.session.rollback()
        flash(f'Lỗi khi cập nhật nhân viên: {e}', 'danger')
    return redirect(url_for('nhan_vien'))

@app.route('/nhan-vien/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_nhan_vien(id):
    emp = Employee.query.get_or_404(id)
    try:
        db.session.delete(emp)
        db.session.commit()
        flash('Đã xóa nhân viên thành công!', 'success')
    except Exception as e:
        print(f"Lỗi xóa nhân viên: {e}")
        db.session.rollback()
        flash(f'Lỗi khi xóa nhân viên: {e}', 'danger')
    return redirect(url_for('nhan_vien'))

@app.route('/khach-hang', methods=['GET', 'POST'])
@login_required
@admin_required
def khach_hang():
    if request.method == 'POST':
        code = request.form['customer_code']
        name = request.form['customer_name']
        
        existing = Customer.query.filter_by(customer_code=code).first()
        if existing:
            flash(f'Mã khách hàng "{code}" đã tồn tại.', 'danger')
            return redirect(url_for('khach_hang'))
            
        try:
            new_cust = Customer(customer_code=code, customer_name=name)
            db.session.add(new_cust)
            db.session.commit()
            flash('Thêm khách hàng thành công!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Lỗi: {e}', 'danger')
        return redirect(url_for('khach_hang'))

    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    per_page = 20
    
    query = Customer.query
    if search:
        query = query.filter(Customer.customer_name.ilike(f'%{search}%') | Customer.customer_code.ilike(f'%{search}%'))
    
    customers = query.order_by(Customer.id.desc()).paginate(page=page, per_page=per_page, error_out=False)
    return render_template('khachhang.html', customers=customers, search_term=search)

@app.route('/khach-hang/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_khach_hang(id):
    cust = Customer.query.get_or_404(id)
    new_code = request.form['customer_code']
    
    existing = Customer.query.filter(Customer.customer_code == new_code, Customer.id != id).first()
    if existing:
        flash(f'Mã "{new_code}" đã tồn tại.', 'danger')
        return redirect(url_for('khach_hang'))
        
    try:
        cust.customer_code = new_code
        cust.customer_name = request.form['customer_name']
        db.session.commit()
        flash('Cập nhật thành công!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi: {e}', 'danger')
    return redirect(url_for('khach_hang'))

@app.route('/khach-hang/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_khach_hang(id):
    cust = Customer.query.get_or_404(id)
    try:
        db.session.delete(cust)
        db.session.commit()
        flash('Xóa thành công!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi: {e}', 'danger')
    return redirect(url_for('khach_hang'))

@app.route('/account', methods=['GET', 'POST'])
@login_required
@admin_required
def account():
    if request.method == 'POST':
        customer_id = request.form.get('customer_id')
        account_code = request.form.get('account_code')
        account_name = request.form.get('account_name')
        is_active = True if request.form.get('is_active') else False
        
        existing = CustomerAccount.query.filter_by(customer_id=customer_id, account_code=account_code).first()
        if existing:
            flash(f'Mã account "{account_code}" đã tồn tại cho khách hàng này.', 'danger')
            return redirect(url_for('account'))
            
        try:
            new_acc = CustomerAccount(customer_id=customer_id, account_code=account_code, account_name=account_name, is_active=is_active)
            db.session.add(new_acc)
            db.session.commit()
            flash('Thêm account thành công!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Lỗi: {e}', 'danger')
        return redirect(url_for('account'))

    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    query = CustomerAccount.query.join(Customer)
    
    if search:
        query = query.filter(CustomerAccount.account_code.ilike(f'%{search}%') | CustomerAccount.account_name.ilike(f'%{search}%') | Customer.customer_name.ilike(f'%{search}%'))
    
    accounts = query.order_by(CustomerAccount.id.desc()).paginate(page=page, per_page=20, error_out=False)
    customers = Customer.query.order_by(Customer.customer_name).all()
    return render_template('account.html', accounts=accounts, customers=customers, search_term=search)

@app.route('/account/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_account(id):
    acc = CustomerAccount.query.get_or_404(id)
    acc.account_code = request.form['account_code']
    acc.account_name = request.form['account_name']
    acc.customer_id = request.form['customer_id']
    acc.is_active = True if request.form.get('is_active') else False
    db.session.commit()
    flash('Cập nhật account thành công!', 'success')
    return redirect(url_for('account'))

@app.route('/account/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_account(id):
    acc = CustomerAccount.query.get_or_404(id)
    db.session.delete(acc)
    db.session.commit()
    flash('Xóa account thành công!', 'success')
    return redirect(url_for('account'))

@app.route('/account-tasks', methods=['GET', 'POST'])
@login_required
@admin_required
def account_tasks():
    if request.method == 'POST':
        account_id = request.form['account_id']
        task_code = request.form['task_code']
        task_name = request.form['task_name']
        
        try:
            new_task = AccountTask(account_id=account_id, task_code=task_code, task_name=task_name)
            db.session.add(new_task)
            db.session.commit()
            flash('Thêm task thành công!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Lỗi: {e}', 'danger')
        return redirect(url_for('account_tasks', account_id=account_id))

    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    account_id = request.args.get('account_id', type=int)
    
    query = AccountTask.query.join(CustomerAccount)
    
    if account_id:
        query = query.filter(AccountTask.account_id == account_id)
    
    if search:
        query = query.filter(AccountTask.task_code.ilike(f'%{search}%') | AccountTask.task_name.ilike(f'%{search}%'))
    
    tasks = query.order_by(AccountTask.id.desc()).paginate(page=page, per_page=20, error_out=False)
    accounts = CustomerAccount.query.filter_by(is_active=True).all()
    selected_account = CustomerAccount.query.get(account_id) if account_id else None
    
    return render_template('account_tasks.html', tasks=tasks, accounts=accounts, search_term=search, selected_account=selected_account)

@app.route('/account-tasks/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_account_task(id):
    task = AccountTask.query.get_or_404(id)
    task.task_code = request.form['task_code']
    task.task_name = request.form['task_name']
    task.account_id = request.form['account_id']
    db.session.commit()
    flash('Cập nhật task thành công!', 'success')
    return redirect(url_for('account_tasks', account_id=task.account_id))

@app.route('/account-tasks/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_account_task(id):
    task = AccountTask.query.get_or_404(id)
    account_id = task.account_id
    db.session.delete(task)
    db.session.commit()
    flash('Xóa task thành công!', 'success')
    return redirect(url_for('account_tasks', account_id=account_id))

@app.route('/account-conversion-index', methods=['GET', 'POST'])
@login_required
@admin_required
def account_conversion_index():
    if request.method == 'POST':
        try:
            account_id = request.form['account_id']
            task_id = request.form['task_id']
            conversion_index = request.form['conversion_index']
            unit = request.form['unit']
            effective_from = datetime.strptime(request.form['effective_from'], '%Y-%m-%d').date()
            effective_to_str = request.form.get('effective_to')
            effective_to = datetime.strptime(effective_to_str, '%Y-%m-%d').date() if effective_to_str else None
            
            new_idx = AccountConversionIndex(
                account_id=account_id,
                task_id=task_id,
                conversion_index=conversion_index,
                unit=unit,
                effective_from=effective_from,
                effective_to=effective_to
            )
            db.session.add(new_idx)
            db.session.commit()
            flash('Thêm định mức thành công!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Lỗi: {e}', 'danger')
        return redirect(url_for('account_conversion_index'))

    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    sort_by = request.args.get('sort_by', 'id')
    order = request.args.get('order', 'desc')
    
    query = AccountConversionIndex.query.join(CustomerAccount).join(Customer).join(AccountTask)
    query = AccountConversionIndex.query.join(CustomerAccount).join(Customer).join(AccountConversionIndex.task)
    if search:
        query = query.filter(
            CustomerAccount.account_code.ilike(f'%{search}%') | 
            CustomerAccount.account_name.ilike(f'%{search}%') | 
            Customer.customer_name.ilike(f'%{search}%') | 
            AccountTask.task_name.ilike(f'%{search}%') |
            AccountTask.task_code.ilike(f'%{search}%')
        )
    
    # Xử lý sắp xếp
    if sort_by == 'account':
        if order == 'asc':
            query = query.order_by(CustomerAccount.account_code.asc())
        else:
            query = query.order_by(CustomerAccount.account_code.desc())
    elif sort_by == 'task':
        if order == 'asc':
            query = query.order_by(AccountTask.task_name.asc())
        else:
            query = query.order_by(AccountTask.task_name.desc())
    else:
        query = query.order_by(AccountConversionIndex.id.desc())

    indices = query.paginate(page=page, per_page=20, error_out=False)
    accounts = CustomerAccount.query.filter_by(is_active=True).order_by(CustomerAccount.account_code).all()
    
    return render_template('account_conversion_index.html', indices=indices, accounts=accounts, search_term=search, sort_by=sort_by, order=order)

@app.route('/account-conversion-index/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_account_conversion_index(id):
    idx = AccountConversionIndex.query.get_or_404(id)
    try:
        idx.account_id = request.form['account_id']
        idx.task_id = request.form['task_id']
        idx.conversion_index = request.form['conversion_index']
        idx.unit = request.form['unit']
        idx.effective_from = datetime.strptime(request.form['effective_from'], '%Y-%m-%d').date()
        effective_to_str = request.form.get('effective_to')
        idx.effective_to = datetime.strptime(effective_to_str, '%Y-%m-%d').date() if effective_to_str else None
        
        db.session.commit()
        flash('Cập nhật định mức thành công!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi: {e}', 'danger')
    return redirect(url_for('account_conversion_index'))

@app.route('/account-conversion-index/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_account_conversion_index(id):
    idx = AccountConversionIndex.query.get_or_404(id)
    try:
        db.session.delete(idx)
        db.session.commit()
        flash('Xóa định mức thành công!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi: {e}', 'danger')
    return redirect(url_for('account_conversion_index'))

@app.route('/api/tasks-by-account/<int:account_id>')
@login_required
@admin_required
def get_tasks_by_account(account_id):
    tasks = AccountTask.query.filter_by(account_id=account_id).all()
    return jsonify([{'id': t.id, 'code': t.task_code, 'name': t.task_name} for t in tasks])

@app.route('/api/next-account-code/<int:customer_id>')
@login_required
@admin_required
def next_account_code(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    
    # Tìm account cuối cùng của khách hàng này
    last_acc = CustomerAccount.query.filter_by(customer_id=customer_id)\
        .order_by(CustomerAccount.id.desc()).first()
    
    # Mặc định: Mã KH + "-01"
    next_code = f"{customer.customer_code}-01"
    
    if last_acc:
        # Tìm phần số ở cuối chuỗi (ví dụ: ACC001 -> 001)
        match = re.search(r'(\d+)$', last_acc.account_code)
        if match:
            number_str = match.group(1)
            number = int(number_str) + 1
            # Giữ nguyên độ dài padding (ví dụ 01 -> 02, 001 -> 002)
            padding = len(number_str)
            prefix = last_acc.account_code[:match.start()]
            next_code = f"{prefix}{str(number).zfill(padding)}"
    
    return jsonify({'next_code': next_code})

@app.route('/import-data', methods=['GET', 'POST'])
@login_required
@update_required
def import_data():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Không tìm thấy file tải lên.', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('Chưa chọn file.', 'danger')
            return redirect(request.url)
            
        if file and file.filename.endswith(('.xlsx', '.xls')):
            try:
                # Đọc file Excel bằng pandas
                # Sử dụng engine openpyxl để đảm bảo đọc tốt file xlsx
                df = pd.read_excel(file, engine='openpyxl')

                # --- 1. ĐỊNH NGHĨA MAPPING CỘT (Hỗ trợ nhiều cách gọi tên) ---
                col_map = {
                    'date': 'date', 'ngày': 'date', 'ngay': 'date', 'work date': 'date', 'ngay lam viec': 'date',
                    'số cont/xe': 'container_no', 'container': 'container_no', 'cont': 'container_no', 'số xe': 'container_no', 'số cont': 'container_no', 'so cont': 'container_no', 'so xe': 'container_no',
                    'task': 'task', 'hạng mục': 'task', 'công việc': 'task', 'cong viec': 'task', 'ten hang muc': 'task',
                    'account': 'account', 'tài khoản': 'account', 'tai khoan': 'account',
                    'khách hàng': 'customer', 'customer': 'customer', 'khach hang': 'customer', 'ten khach hang': 'customer',
                    'cbm': 'cbm', 'sản lượng': 'cbm', 'sl': 'cbm', 'khối lượng': 'cbm', 'san luong': 'cbm', 'khoi luong': 'cbm',
                    'tally': 'tally', 'kiểm đếm': 'tally', 'kiem dem': 'tally',
                    'xe nang': 'lift_truck', 'xe nâng': 'lift_truck', 'lift truck': 'lift_truck', 'lai xe nang': 'lift_truck',
                    'cong nhan_1': 'worker_1', 'cong nhan 1': 'worker_1', 'worker 1': 'worker_1', 'công nhân 1': 'worker_1',
                    'cong nhan_2': 'worker_2', 'cong nhan 2': 'worker_2', 'worker 2': 'worker_2', 'công nhân 2': 'worker_2',
                    'cong nhan_3': 'worker_3', 'cong nhan 3': 'worker_3', 'worker 3': 'worker_3', 'công nhân 3': 'worker_3',
                    'cong nhan_4': 'worker_4', 'cong nhan 4': 'worker_4', 'worker 4': 'worker_4', 'công nhân 4': 'worker_4',
                    'cong nhan_5': 'worker_5', 'cong nhan 5': 'worker_5', 'worker 5': 'worker_5', 'công nhân 5': 'worker_5',
                    'cong nhan_6': 'worker_6', 'cong nhan 6': 'worker_6', 'worker 6': 'worker_6', 'công nhân 6': 'worker_6',
                }

                # Hàm chuẩn hóa tên cột (về chữ thường, bỏ dấu, bỏ khoảng trắng thừa)
                def normalize_str(s):
                    if pd.isna(s): return ""
                    return unicodedata.normalize('NFC', str(s)).strip().lower()

                # --- 2. TỰ ĐỘNG TÌM DÒNG TIÊU ĐỀ ---
                # Nếu dòng đầu tiên không chứa đủ từ khóa quan trọng, quét 20 dòng đầu
                current_cols = [normalize_str(c) for c in df.columns]
                # Đếm số lượng cột khớp với mapping
                match_count = sum(1 for c in current_cols if c in col_map)
                
                # Nếu ít hơn 3 cột khớp, coi như chưa tìm thấy header đúng -> Quét lại
                if match_count < 3:
                    file.seek(0)
                    # Đọc header=None để lấy dữ liệu thô
                    df_raw = pd.read_excel(file, header=None, nrows=20, engine='openpyxl') 
                    found_header_idx = -1
                    
                    for idx, row in df_raw.iterrows():
                        row_vals = [normalize_str(v) for v in row.astype(str)]
                        # Kiểm tra dòng này có bao nhiêu từ khóa khớp
                        matches = sum(1 for v in row_vals if v in col_map)
                        if matches >= 3: # Ngưỡng chấp nhận: tìm thấy ít nhất 3 cột quen thuộc
                            found_header_idx = idx
                            break
                    
                    if found_header_idx != -1:
                        file.seek(0)
                        df = pd.read_excel(file, header=found_header_idx, engine='openpyxl')

                # --- 3. ĐỔI TÊN CỘT VỀ CHUẨN (Standard Keys) ---
                new_columns = []
                for col in df.columns:
                    norm_col = normalize_str(col)
                    # Nếu tên cột nằm trong map thì lấy tên chuẩn, không thì giữ nguyên
                    new_columns.append(col_map.get(norm_col, col))
                df.columns = new_columns

                # --- 4. KIỂM TRA CỘT BẮT BUỘC (Dùng tên chuẩn) ---
                required_keys = ['date', 'container_no', 'task', 'account', 'customer']
                missing = [k for k in required_keys if k not in df.columns]

                if missing:
                    flash(f'Lỗi file: Không tìm thấy các cột bắt buộc: {", ".join(missing)}.<br>Vui lòng kiểm tra lại tên cột trong file Excel.', 'danger')
                    return redirect(request.url)
                
                # Xử lý cột Date: chuyển đổi chuỗi sang datetime (hỗ trợ DD/MM/YYYY)
                if 'date' in df.columns:
                    df['date'] = pd.to_datetime(df['date'], dayfirst=True, errors='coerce')

                # Xử lý NaN thành None để tránh lỗi DB (chuyển sang object để giữ None)
                df = df.where(pd.notnull(df), None)
                
                # Xóa dữ liệu tạm cũ trước khi import mới
                db.session.query(LaborProductivityTemp).delete()
                
                # --- TỐI ƯU HÓA: CHUẨN BỊ DỮ LIỆU ĐỂ BULK INSERT ---
                bulk_data = []
                for row in df.to_dict('records'):
                    # Xử lý ngày tháng
                    d = row.get('date')
                    # Kiểm tra kỹ hơn để tránh lỗi NaT (Not a Time) của pandas
                    if pd.isnull(d):
                        d = None
                    elif hasattr(d, 'date'):
                        d = d.date()
                    
                    # Xử lý an toàn cho CBM (tránh lỗi nếu file excel có chữ trong cột số)
                    raw_cbm = row.get('cbm')
                    safe_cbm = None
                    if raw_cbm is not None:
                        try:
                            safe_cbm = float(raw_cbm)
                        except (ValueError, TypeError):
                            safe_cbm = 0.0

                    # Helper để lấy string sạch (cắt khoảng trắng thừa)
                    def get_str(key):
                        val = row.get(key)
                        return str(val).strip() if val is not None else None

                    bulk_data.append({
                        'date': d,
                        'container_no': get_str('container_no'),
                        'cbm': safe_cbm,
                        'tally': get_str('tally'),
                        'lift_truck': get_str('lift_truck'),
                        'worker_1': get_str('worker_1'),
                        'worker_2': get_str('worker_2'),
                        'worker_3': get_str('worker_3'),
                        'worker_4': get_str('worker_4'),
                        'worker_5': get_str('worker_5'),
                        'worker_6': get_str('worker_6'),
                        'task': get_str('task'),
                        'account': get_str('account'),
                        'customer': get_str('customer')
                    })
                
                # Insert hàng loạt (Nhanh hơn gấp nhiều lần so với loop add)
                if bulk_data:
                    db.session.bulk_insert_mappings(LaborProductivityTemp, bulk_data)
                    db.session.commit()
                    flash(f'Đã đọc {len(df)} dòng vào bảng tạm. Vui lòng kiểm tra và xác nhận lưu!', 'success')
                else:
                    flash('File không có dữ liệu.', 'warning')

            except Exception as e:
                db.session.rollback()
                flash(f'Lỗi khi đọc file: {str(e)}', 'danger')
        else:
            flash('Vui lòng chỉ tải lên file Excel (.xlsx, .xls)', 'danger')
            
    # Lấy dữ liệu tạm (nếu có)
    temp_records = LaborProductivityTemp.query.order_by(LaborProductivityTemp.id).all()
    preview_data = []
    has_errors = False

    if temp_records:
        # Tạo một map để kiểm tra hiệu quả: { 'customer_name_lower': {'account_name_lower', ...} }
        customer_accounts_map = {}
        all_accounts = CustomerAccount.query.join(Customer).with_entities(Customer.customer_name, CustomerAccount.account_name).all()
        for cust_name, acc_name in all_accounts:
            cust_key = cust_name.strip().lower()
            if cust_key not in customer_accounts_map:
                customer_accounts_map[cust_key] = set()
            if acc_name:
                customer_accounts_map[cust_key].add(acc_name.strip().lower())

        for t in temp_records:
            is_row_valid = True
            cust_name = t.customer.strip().lower() if t.customer else ''
            acc_name = t.account.strip().lower() if t.account else ''

            # Kiểm tra xem khách hàng có tồn tại và account có thuộc khách hàng đó không
            if not cust_name or not acc_name or cust_name not in customer_accounts_map or acc_name not in customer_accounts_map.get(cust_name, set()):
                is_row_valid = False
                has_errors = True
            
            preview_data.append({'record': t, 'is_valid': is_row_valid})
    
    # Lấy dữ liệu chính thức để hiển thị (phân trang)
    page = request.args.get('page', 1, type=int)
    records = LaborProductivity.query.order_by(LaborProductivity.id.desc()).paginate(page=page, per_page=20, error_out=False)
    
    return render_template('importdata.html', records=records, preview_data=preview_data, has_errors=has_errors)

@app.route('/import-data/confirm', methods=['POST'])
@login_required
@update_required
def confirm_import():
    try:
        temps = LaborProductivityTemp.query.order_by(LaborProductivityTemp.id).all()
        if not temps:
            flash('Không có dữ liệu tạm để lưu.', 'warning')
            return redirect(url_for('import_data'))
            
        # --- TỐI ƯU HÓA: TẢI TRƯỚC DỮ LIỆU VÀO RAM (CACHE) ---
        # Thay vì query trong vòng lặp, ta query 1 lần và lưu vào Dictionary
        
        # 1. Cache Customers: { 'ten_kh_lower': id }
        customers_map = {c.customer_name.strip().lower(): c for c in Customer.query.all()}
        
        # 2. Cache Accounts: { (customer_id, 'ten_acc_lower'): account_obj }
        accounts_map = {}
        for acc in CustomerAccount.query.all():
            key = (acc.customer_id, acc.account_name.strip().lower())
            accounts_map[key] = acc

        # 3. Cache Tasks: { (account_id, 'code_or_name_lower'): task_obj }
        tasks_map = {}
        for t in AccountTask.query.all():
            # Map cả code và name để tìm kiếm linh hoạt
            tasks_map[(t.account_id, t.task_code.strip().lower())] = t
            tasks_map[(t.account_id, t.task_name.strip().lower())] = t

        # 4. Cache Conversion Indices: { (account_id, task_id): index_obj }
        # Lấy tất cả index, sắp xếp theo ngày hiệu lực tăng dần
        # Khi đưa vào dict, giá trị sau sẽ ghi đè giá trị trước -> Lấy được cái mới nhất
        indices_map = {}
        all_indices = AccountConversionIndex.query.order_by(AccountConversionIndex.effective_from).all()
        for idx in all_indices:
            indices_map[(idx.account_id, idx.task_id)] = idx

        # Danh sách chứa dữ liệu để insert hàng loạt
        bulk_insert_list = []
        
        # --- BƯỚC 1: VALIDATE DỮ LIỆU TRƯỚC KHI LƯU ---
        errors = []
        for i, t in enumerate(temps):
            if not t.customer or not t.account:
                errors.append(f"Dòng {i + 1}: Thiếu thông tin Khách hàng hoặc Account.")
                continue

            cust_key = t.customer.strip().lower()
            acc_key = t.account.strip().lower()
            
            customer = customers_map.get(cust_key)
            if not customer:
                errors.append(f"Dòng {i + 1}: Khách hàng '{t.customer}' không tồn tại trong hệ thống.")
                continue

            acc = accounts_map.get((customer.id, acc_key))
            if not acc:
                errors.append(f"Dòng {i + 1}: Account '{t.account}' không thuộc khách hàng '{t.customer}'.")

        if errors:
            flash('Không thể lưu do có lỗi dữ liệu. Vui lòng kiểm tra lại:', 'danger')
            for error in errors:
                flash(error, 'danger')
            return redirect(url_for('import_data'))

        # --- BƯỚC 2: LƯU DỮ LIỆU NẾU KHÔNG CÓ LỖI ---
        for t in temps:
            # Mặc định ban đầu
            conv_index = 1.0
            unit = 'CBM'
            quantity = t.cbm if t.cbm is not None else 0.0
            
            # Lấy object từ Cache (đã validate ở trên nên chắc chắn có Customer và Account)
            cust_key = t.customer.strip().lower()
            customer = customers_map[cust_key]
            
            acc_key = t.account.strip().lower()
            acc = accounts_map[(customer.id, acc_key)]
            
            task_obj = None

            # Tìm định mức chuyển đổi dựa trên Account Code và Task Code
            if t.task:
                task_val = t.task.strip().lower()
                # Tìm Task trong Cache
                task_obj = tasks_map.get((acc.id, task_val))

                if task_obj:
                    # Tìm Index trong Cache
                    idx_index = indices_map.get((acc.id, task_obj.id))
                    
                    if idx_index:
                        conv_index = float(idx_index.conversion_index)
                        unit = idx_index.unit
                        if t.cbm is not None:
                            quantity = float(t.cbm) * conv_index
            
            # Sử dụng tên từ các object đã tìm thấy để lưu, nếu không tìm thấy thì dùng tên gốc từ Excel.
            task_name_to_save = task_obj.task_name if task_obj else t.task
            account_name_to_save = acc.account_name if acc else t.account
            customer_name_to_save = customer.customer_name if customer else t.customer

            bulk_insert_list.append({
                'work_date': t.date,
                'ref_no': t.container_no,
                'productivity_value': t.cbm,
                'tally_id': t.tally,
                'xenang_id': t.lift_truck,
                'congnhan1_id': t.worker_1,
                'congnhan2_id': t.worker_2,
                'congnhan3_id': t.worker_3,
                'congnhan4_id': t.worker_4,
                'congnhan5_id': t.worker_5,
                'congnhan6_id': t.worker_6,
                'task_id': task_name_to_save,
                'account_id': account_name_to_save,
                'customer_id': customer_name_to_save,
                'unit': unit,
                'conversion_index': conv_index,
                'quantity': quantity
            })
        
        # Insert hàng loạt vào bảng chính
        if bulk_insert_list:
            db.session.bulk_insert_mappings(LaborProductivity, bulk_insert_list)
            
        # Xóa dữ liệu tạm sau khi lưu thành công
        db.session.query(LaborProductivityTemp).delete()
        db.session.commit()
        flash(f'Đã lưu chính thức {len(bulk_insert_list)} dòng dữ liệu!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi khi lưu dữ liệu: {str(e)}', 'danger')
    return redirect(url_for('import_data'))

@app.route('/import-data/cancel', methods=['POST'])
@login_required
@update_required
def cancel_import():
    try:
        db.session.query(LaborProductivityTemp).delete()
        db.session.commit()
        flash('Đã hủy bỏ dữ liệu tạm.', 'info')
    except Exception as e:
        flash(f'Lỗi: {str(e)}', 'danger')
    return redirect(url_for('import_data'))

@app.route('/import-data/update-temp/<int:id>', methods=['POST'])
@login_required
@update_required
def update_temp_data(id):
    try:
        temp = LaborProductivityTemp.query.get_or_404(id)
        
        # Cập nhật dữ liệu từ form
        date_str = request.form.get('date')
        if date_str:
            temp.date = datetime.strptime(date_str, '%Y-%m-%d').date()
            
        temp.container_no = request.form.get('container_no')
        temp.cbm = request.form.get('cbm')
        temp.tally = request.form.get('tally')
        temp.lift_truck = request.form.get('lift_truck')
        temp.worker_1 = request.form.get('worker_1')
        temp.worker_2 = request.form.get('worker_2')
        temp.worker_3 = request.form.get('worker_3')
        temp.worker_4 = request.form.get('worker_4')
        temp.worker_5 = request.form.get('worker_5')
        temp.worker_6 = request.form.get('worker_6')
        temp.task = request.form.get('task')
        temp.account = request.form.get('account')
        temp.customer = request.form.get('customer')
        
        db.session.commit()
        flash('Cập nhật dòng dữ liệu tạm thành công!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi cập nhật: {str(e)}', 'danger')
    return redirect(url_for('import_data'))

@app.route('/import-data/delete-temp/<int:id>', methods=['POST'])
@login_required
@update_required
def delete_temp_data(id):
    try:
        temp = LaborProductivityTemp.query.get_or_404(id)
        db.session.delete(temp)
        db.session.commit()
        flash('Đã xóa dòng dữ liệu tạm.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi xóa: {str(e)}', 'danger')
    return redirect(url_for('import_data'))

@app.route('/settings', methods=['GET', 'POST'])
@login_required
@admin_required
def settings():
    if request.method == 'POST':
        # Lấy chuỗi nhập vào, ví dụ: "TB, IF, HB"
        prefixes_input = request.form.get('exclusion_prefixes', '')
        
        # Chuẩn hóa: tách dấu phẩy, xóa khoảng trắng, chuyển thành chữ hoa
        # Kết quả lưu DB dạng: "TB,IF,HB"
        prefixes_clean = ",".join([p.strip().upper() for p in prefixes_input.split(',') if p.strip()])
        
        setting = SystemSetting.query.filter_by(key_name='exclusion_prefixes').first()
        if not setting:
            setting = SystemSetting(key_name='exclusion_prefixes', value=prefixes_clean)
            db.session.add(setting)
        else:
            setting.value = prefixes_clean
        
        db.session.commit()
        flash('Cập nhật cài đặt thành công!', 'success')
        return redirect(url_for('settings'))

    setting = SystemSetting.query.filter_by(key_name='exclusion_prefixes').first()
    current_prefixes = setting.value if setting else "TB, IF, HB" # Giá trị mặc định nếu chưa cấu hình
    return render_template('settings.html', current_prefixes=current_prefixes)

@app.route('/users', methods=['GET', 'POST'])
@login_required
@admin_required
def manage_users():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        full_name = request.form['full_name']
        role = request.form['role']
        is_active = True if request.form.get('is_active') else False
        can_export = True if request.form.get('can_export') else False

        existing = User.query.filter_by(username=username).first()
        if existing:
            flash(f'Username "{username}" đã tồn tại.', 'danger')
            return redirect(url_for('manage_users'))
        
        try:
            hashed_pw = generate_password_hash(password)
            new_user = User(username=username, password_hash=hashed_pw, full_name=full_name, role=role, is_active=is_active, can_export=can_export)
            db.session.add(new_user)
            db.session.commit()
            flash('Thêm user thành công!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Lỗi: {e}', 'danger')
        return redirect(url_for('manage_users'))

    page = request.args.get('page', 1, type=int)
    users = User.query.order_by(User.id.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template('users.html', users=users)

@app.route('/users/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_user(id):
    user = User.query.get_or_404(id)

    # Cập nhật các trường chung
    user.full_name = request.form['full_name']
    user.can_export = True if request.form.get('can_export') else False

    # Nếu là user 'admin', không cho phép thay đổi quyền và trạng thái
    if user.username == 'admin':
        user.role = 'ADMIN'
        user.is_active = True
    else:
        # Các user khác thì cập nhật bình thường
        user.role = request.form['role']
        user.is_active = True if request.form.get('is_active') else False

    # Chỉ cập nhật mật khẩu nếu người dùng nhập mới
    new_password = request.form.get('password')
    if new_password:
        user.password_hash = generate_password_hash(new_password)

    try:
        db.session.commit()
        flash('Cập nhật user thành công!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi: {e}', 'danger')
    return redirect(url_for('manage_users'))

@app.route('/users/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_user(id):
    user = User.query.get_or_404(id)
    if user.username == 'admin':
        flash('Không thể xóa tài khoản Admin mặc định.', 'danger')
        return redirect(url_for('manage_users'))

    db.session.delete(user)
    db.session.commit()
    flash('Xóa user thành công!', 'success')
    return redirect(url_for('manage_users'))

@app.route('/api/tasks-by-account-name')
@login_required
def get_tasks_by_account_name():
    account_name = request.args.get('account_name')
    customer_name = request.args.get('customer_name')

    if not account_name or not customer_name:
        return jsonify([])

    customer = Customer.query.filter(Customer.customer_name.ilike(customer_name)).first()
    if not customer:
        return jsonify([])
    
    account = CustomerAccount.query.filter(
        CustomerAccount.customer_id == customer.id,
        CustomerAccount.account_name.ilike(account_name)
    ).first()

    if not account:
        return jsonify([])

    tasks = AccountTask.query.filter_by(account_id=account.id).order_by(AccountTask.task_name).all()
    return jsonify([{'name': t.task_name} for t in tasks])

@app.route('/api/get-conversion-info')
@login_required
def get_conversion_info():
    customer_name = request.args.get('customer_name')
    account_name = request.args.get('account_name')
    task_name = request.args.get('task_name')

    default_response = jsonify({'conversion_index': 1.0, 'unit': 'CBM'})

    if not all([customer_name, account_name, task_name]):
        return default_response

    try:
        customer = Customer.query.filter(Customer.customer_name.ilike(customer_name)).first()
        if not customer: return default_response

        account = CustomerAccount.query.filter(
            CustomerAccount.customer_id == customer.id,
            CustomerAccount.account_name.ilike(account_name)
        ).first()
        if not account: return default_response

        task = AccountTask.query.filter(
            AccountTask.account_id == account.id,
            or_(AccountTask.task_code.ilike(task_name), AccountTask.task_name.ilike(task_name))
        ).first()
        if not task: return default_response

        # Lấy index mới nhất, bỏ qua ngày hiệu lực để nhất quán với logic import
        index = AccountConversionIndex.query.filter(
            AccountConversionIndex.account_id == account.id,
            AccountConversionIndex.task_id == task.id
        ).order_by(AccountConversionIndex.effective_from.desc()).first()

        if index:
            return jsonify({'conversion_index': float(index.conversion_index), 'unit': index.unit})
        else:
            return default_response
    except Exception:
        return default_response

@app.route('/productivity', methods=['GET', 'POST'])
@login_required
@admin_required
def manage_productivity():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    
    query = LaborProductivity.query

    if from_date:
        query = query.filter(LaborProductivity.work_date >= from_date)
    if to_date:
        query = query.filter(LaborProductivity.work_date <= to_date)

    if search:
        query = query.filter(
            LaborProductivity.ref_no.ilike(f'%{search}%') |
            LaborProductivity.task_id.ilike(f'%{search}%') |
            LaborProductivity.account_id.ilike(f'%{search}%') |
            LaborProductivity.customer_id.ilike(f'%{search}%')
        )
    
    # Sắp xếp theo ngày giảm dần, sau đó đến ID giảm dần
    records = query.order_by(LaborProductivity.work_date.desc(), LaborProductivity.id.desc()).paginate(page=page, per_page=20, error_out=False)
    
    return render_template('productivity.html', records=records, search_term=search, from_date=from_date, to_date=to_date)

@app.route('/productivity/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_productivity(id):
    record = LaborProductivity.query.get_or_404(id)
    try:
        record.work_date = datetime.strptime(request.form['work_date'], '%Y-%m-%d').date()
        record.ref_no = request.form['ref_no']
        record.customer_id = request.form['customer_id']
        record.account_id = request.form['account_id']
        record.task_id = request.form['task_id']
        record.quantity = float(request.form['quantity'])
        record.unit = request.form['unit']
        # Cập nhật CBM gốc nếu cần (productivity_value)
        if request.form.get('productivity_value'):
            record.productivity_value = float(request.form['productivity_value'])
            
        db.session.commit()
        flash('Cập nhật sản lượng thành công!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi cập nhật: {e}', 'danger')
    return redirect(url_for('manage_productivity'))

@app.route('/productivity/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_productivity(id):
    record = LaborProductivity.query.get_or_404(id)
    try:
        db.session.delete(record)
        db.session.commit()
        flash('Xóa bản ghi thành công!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi xóa: {e}', 'danger')
    return redirect(url_for('manage_productivity'))

@app.route('/report', methods=['GET', 'POST'])
@login_required
@view_required
def report():
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    
    # Nếu không chọn ngày, mặc định từ 26 tháng trước đến 25 tháng hiện tại
    if not from_date and not to_date:
        today = datetime.now()
        if today.month == 1:
            prev_month = 12
            prev_year = today.year - 1
        else:
            prev_month = today.month - 1
            prev_year = today.year
            
        from_date = f"{prev_year}-{prev_month:02d}-26"
        to_date = f"{today.year}-{today.month:02d}-25"

    query = LaborProductivity.query
    
    if from_date:
        query = query.filter(LaborProductivity.work_date >= from_date)
    if to_date:
        query = query.filter(LaborProductivity.work_date <= to_date)
        
    # Lấy dữ liệu chi tiết
    records = query.order_by(LaborProductivity.work_date.desc()).all()
    
    # --- TỔNG HỢP DỮ LIỆU THEO NHÂN VIÊN ---
    # Dictionary để lưu tổng hợp: { 'Tên NV': {'role': 'Vai trò', 'total_qty': 0.0, 'count': 0} }
    staff_stats = {}
    
    # Lấy cấu hình loại trừ từ DB
    setting = SystemSetting.query.filter_by(key_name='exclusion_prefixes').first()
    # Chuyển chuỗi "TB,IF,HB" thành tuple ('TB', 'IF', 'HB') để dùng cho startswith
    prefixes_str = setting.value if setting else "TB,IF,HB"
    exclusion_prefixes = tuple(p.strip() for p in prefixes_str.split(',') if p.strip())

    # Lấy danh sách mã nhân viên hợp lệ từ bảng Employee để kiểm tra
    valid_codes = set(r[0] for r in db.session.query(Employee.masl).filter(Employee.masl != None, Employee.masl != "").all())

    def add_stat(name, role, qty):
        if not name: return
        
        # Bỏ qua các mã bắt đầu bằng prefix cấu hình
        if str(name).upper().startswith(exclusion_prefixes):
            return
            
        # Tạo key duy nhất là (Tên, Vai trò) để phân biệt nếu 1 người làm nhiều vai trò
        key = (name, role)
        if key not in staff_stats:
            # Kiểm tra xem mã nhân viên có trong danh sách không
            is_valid = name in valid_codes
            remark = "" if is_valid else "Không có trong danh sách"
            
            staff_stats[key] = {'name': name, 'role': role, 'total_qty': 0.0, 'count': 0, 'remark': remark}
        
        staff_stats[key]['total_qty'] += (qty or 0.0)
        staff_stats[key]['count'] += 1

    for r in records:
        qty = r.quantity if r.quantity is not None else 0.0
        
        # Cộng dồn cho từng vị trí
        add_stat(r.tally_id, 'Tally', qty)
        add_stat(r.xenang_id, 'Xe nâng', qty)
        add_stat(r.congnhan1_id, 'Công nhân', qty)
        add_stat(r.congnhan2_id, 'Công nhân', qty)
        add_stat(r.congnhan3_id, 'Công nhân', qty)
        add_stat(r.congnhan4_id, 'Công nhân', qty)
        add_stat(r.congnhan5_id, 'Công nhân', qty)
        add_stat(r.congnhan6_id, 'Công nhân', qty)
    
    # Chuyển đổi sang list để hiển thị và sắp xếp
    summary = list(staff_stats.values())
    
    # Lấy Top 5 nhân viên có năng suất cao nhất cho biểu đồ
    top_employees = sorted(summary, key=lambda x: x['total_qty'], reverse=True)[:5]
    
    # Sắp xếp danh sách hiển thị bảng theo tên (A-Z)
    summary.sort(key=lambda x: x['name'])
    
    # --- TỔNG HỢP THEO KHÁCH HÀNG ---
    customer_stats = {}
    for r in records:
        qty = r.quantity if r.quantity is not None else 0.0
        c_name = r.customer_id if r.customer_id else "Khác"
        
        if c_name not in customer_stats:
            customer_stats[c_name] = {'name': c_name, 'total_qty': 0.0, 'count': 0}
        customer_stats[c_name]['total_qty'] += qty
        customer_stats[c_name]['count'] += 1
        
    customer_summary = list(customer_stats.values())
    customer_summary.sort(key=lambda x: x['total_qty'], reverse=True)
    
    # --- TỔNG HỢP CHO NHÂN VIÊN AN_CHUNG ---
    an_chung_data = []
    an_chung_emps = Employee.query.filter_by(employee_type='An_chung').all()
    
    # Hàm chuẩn hóa chuỗi để so sánh chính xác hơn
    def normalize_key(s):
        if not s: return None
        return unicodedata.normalize('NFC', str(s)).strip().lower()

    # Map để tra cứu nhanh: key -> Employee Object
    ac_map = {}
    # Map để tổng hợp số liệu (bao gồm cả nhân viên chưa có sản lượng)
    summary_map = {}

    for emp in an_chung_emps:
        if emp.masl: ac_map[normalize_key(emp.masl)] = emp
        if emp.employee_code: ac_map[normalize_key(emp.employee_code)] = emp
        if emp.full_name: ac_map[normalize_key(emp.full_name)] = emp
        
        # Khởi tạo dữ liệu tổng hợp cho TẤT CẢ nhân viên An Chung
        summary_map[emp.employee_code] = {
            'employee_code': emp.employee_code,
            'masl': emp.masl,
            'full_name': emp.full_name,
            'position': emp.position,
            'total_productivity': 0.0,
            'total_quantity': 0.0,
            'count': 0
        }

    for r in records:
        workers = [
            r.tally_id, r.xenang_id, 
            r.congnhan1_id, r.congnhan2_id, r.congnhan3_id, 
            r.congnhan4_id, r.congnhan5_id, r.congnhan6_id
        ]
        seen_in_row = set()
        for w_str in workers:
            if not w_str: continue
            w_key = normalize_key(w_str)
            if w_key in ac_map:
                emp = ac_map[w_key]
                if emp.id in seen_in_row: continue
                seen_in_row.add(emp.id)
                
                an_chung_data.append({
                    'work_date': r.work_date,
                    'employee_code': emp.employee_code,
                    'masl': emp.masl,
                    'full_name': emp.full_name,
                    'position': emp.position,
                    'task_id': r.task_id,
                    'productivity_value': r.productivity_value,
                    'conversion_index': r.conversion_index,
                    'quantity': r.quantity
                })

    # Cập nhật số liệu từ dữ liệu chi tiết vào bảng tổng hợp
    for item in an_chung_data:
        code = item['employee_code']
        if code in summary_map:
            
            p_val = item['productivity_value'] if item['productivity_value'] is not None else 0.0
            q_val = item['quantity'] if item['quantity'] is not None else 0.0
            
            summary_map[code]['total_productivity'] += p_val
            summary_map[code]['total_quantity'] += q_val
            summary_map[code]['count'] += 1
        
    an_chung_summary_list = list(summary_map.values())
    an_chung_summary_list.sort(key=lambda x: x['full_name'])

    return render_template('report.html', records=records, summary=summary, top_employees=top_employees, customer_summary=customer_summary, an_chung_data=an_chung_data, an_chung_summary_list=an_chung_summary_list, from_date=from_date, to_date=to_date)

@app.route('/report/export')
@login_required
@view_required
def export_report():
    if not current_user.can_export:
        flash('B?n kh?ng c? quy?n xu?t b?o c?o.', 'danger')
        return redirect(url_for('report'))

    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')

    query = LaborProductivity.query
    if from_date:
        query = query.filter(LaborProductivity.work_date >= from_date)
    if to_date:
        query = query.filter(LaborProductivity.work_date <= to_date)

    records = query.order_by(LaborProductivity.work_date.desc()).all()

    def normalize_key(s):
        if not s:
            return None
        return unicodedata.normalize('NFC', str(s)).strip().lower()

    palette = [
        'FF00B050', 'FF00B0F0', 'FF70AD47', 'FF92D050',
        'FFFFC000', 'FFED7D31', 'FFC00000', 'FF7030A0'
    ]

    # Lấy index mới nhất theo account_id (ưu tiên effective_from mới nhất, sau đó id mới nhất)
    latest_index_by_account = {}
    idx_rows = AccountConversionIndex.query.order_by(
        AccountConversionIndex.account_id.asc(),
        AccountConversionIndex.effective_from.desc(),
        AccountConversionIndex.id.desc(),
    ).all()
    for idx in idx_rows:
        if idx.account_id not in latest_index_by_account:
            latest_index_by_account[idx.account_id] = float(idx.conversion_index or 1.0)

    account_configs = []
    seen_account_names = set()
    if latest_index_by_account:
        accounts = CustomerAccount.query.filter(
            CustomerAccount.id.in_(latest_index_by_account.keys())
        ).order_by(CustomerAccount.account_name.asc()).all()

        for i, acc in enumerate(accounts):
            account_name = (acc.account_name or '').strip()
            if not account_name:
                continue
            name_key = normalize_key(account_name)
            if name_key in seen_account_names:
                continue
            seen_account_names.add(name_key)
            account_configs.append({
                'title': account_name,
                'coef': latest_index_by_account.get(acc.id, 1.0),
                'color': palette[i % len(palette)],
            })

    def is_an_chung(employee_type):
        t = normalize_key(employee_type) or ''
        return t in ('an_chung', 'an chung')

    all_emps = Employee.query.all()
    an_chung_emps = [e for e in all_emps if is_an_chung(e.employee_type)]
    khoan_emps = [e for e in all_emps if not is_an_chung(e.employee_type)]

    def build_maps(employees):
        worker_map = {}
        summary_map = {}
        for emp in employees:
            if emp.masl:
                worker_map[normalize_key(emp.masl)] = emp
            if emp.employee_code:
                worker_map[normalize_key(emp.employee_code)] = emp
            if emp.full_name:
                worker_map[normalize_key(emp.full_name)] = emp

            row_key = emp.employee_code or f"EMP_{emp.id}"
            summary_map[row_key] = {
                'employee_code': emp.employee_code or '',
                'masl': emp.masl or '',
                'full_name': emp.full_name or '',
                'position': emp.position or '',
                'total_raw': 0.0,
                'total_converted': 0.0,
            }
            for cfg in account_configs:
                summary_map[row_key][cfg['title']] = 0.0
        return worker_map, summary_map

    khoan_map, khoan_summary_map = build_maps(khoan_emps)
    ac_map, an_chung_summary_map = build_maps(an_chung_emps)

    an_chung_detail_data = []
    for r in records:
        workers = [
            r.tally_id,
            r.xenang_id,
            r.congnhan1_id,
            r.congnhan2_id,
            r.congnhan3_id,
            r.congnhan4_id,
            r.congnhan5_id,
            r.congnhan6_id,
        ]

        raw_cbm = float(r.productivity_value or 0.0)
        converted_cbm = float(r.quantity or 0.0)
        matched_account = None
        record_account_key = normalize_key(r.account_id)
        if record_account_key:
            for cfg in account_configs:
                if normalize_key(cfg['title']) == record_account_key:
                    matched_account = cfg['title']
                    break

        seen_in_row = set()
        for worker in workers:
            worker_key = normalize_key(worker)
            if not worker_key:
                continue

            group = None
            emp = None
            summary_map = None
            if worker_key in khoan_map:
                group = 'khoan'
                emp = khoan_map[worker_key]
                summary_map = khoan_summary_map
            elif worker_key in ac_map:
                group = 'an_chung'
                emp = ac_map[worker_key]
                summary_map = an_chung_summary_map
            else:
                continue

            if emp.id in seen_in_row:
                continue
            seen_in_row.add(emp.id)

            row_key = emp.employee_code or f"EMP_{emp.id}"
            if row_key not in summary_map:
                continue

            summary_map[row_key]['total_raw'] += raw_cbm
            summary_map[row_key]['total_converted'] += converted_cbm
            if matched_account:
                summary_map[row_key][matched_account] += raw_cbm

            if group == 'an_chung':
                an_chung_detail_data.append({
                    'Ng?y': r.work_date,
                    'M? NV': emp.employee_code,
                    'MS': emp.masl,
                    'H? v? t?n': emp.full_name,
                    'V? tr?': emp.position,
                    'Task': r.task_id,
                    'Account': r.account_id,
                    'Kh?ch h?ng': r.customer_id,
                    'CBM ch?a h? s?': raw_cbm,
                    'CBM c? h? s?': converted_cbm,
                })

    summary_rows_khoan = list(khoan_summary_map.values())
    summary_rows_khoan.sort(key=lambda x: (x['full_name'] or '').lower())
    summary_rows_an_chung = list(an_chung_summary_map.values())
    summary_rows_an_chung.sort(key=lambda x: (x['full_name'] or '').lower())

    def make_summary_df(summary_rows):
        return pd.DataFrame([
            {
                'STT': idx,
                'VTCV': item['position'],
                'MSNV': item['employee_code'],
                'MS': item['masl'],
                'H? V? T?N': item['full_name'],
                'T?ng CBM C? H? S?': item['total_converted'],
                'T?ng CBM CH?A H? S?': item['total_raw'],
                **{cfg['title']: item[cfg['title']] for cfg in account_configs},
            }
            for idx, item in enumerate(summary_rows, 1)
        ])

    df_summary_template_khoan = make_summary_df(summary_rows_khoan)
    df_summary_template_an_chung = make_summary_df(summary_rows_an_chung)

    detail_data = []
    for idx, r in enumerate(records, 1):
        workers = [r.congnhan1_id, r.congnhan2_id, r.congnhan3_id, r.congnhan4_id, r.congnhan5_id, r.congnhan6_id]
        has_worker = any(w for w in workers if w)
        cbm_val = r.productivity_value if has_worker else ''

        detail_data.append({
            'STT': idx,
            'Ng?y nh?p h?ng': r.work_date,
            'S? xe/cont': r.ref_no,
            'CBM': cbm_val,
            'Quantity': r.quantity,
            'Tally': r.tally_id,
            'Xe N?ng': r.xenang_id,
            'C?ng nh?n 1': r.congnhan1_id,
            'C?ng nh?n 2': r.congnhan2_id,
            'C?ng nh?n 3': r.congnhan3_id,
            'C?ng nh?n 4': r.congnhan4_id,
            'C?ng nh?n 5': r.congnhan5_id,
            'C?ng nh?n 6': r.congnhan6_id,
            'Task': r.task_id,
            'Account': r.account_id,
            'Kh?ch h?ng': r.customer_id,
        })

    df_detail = pd.DataFrame(detail_data)
    df_anchung = pd.DataFrame(an_chung_detail_data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        wb = writer.book
        thin = Side(style='thin', color='000000')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        fill_header_left = PatternFill(fill_type='solid', fgColor='FFDAF2D0')
        fill_yellow = PatternFill(fill_type='solid', fgColor='FFFFFF00')
        total_data_cols = 7 + len(account_configs)
        base_headers = ['STT', 'VTCV', 'MSNV', 'MS (NEU CO)', 'HO VA TEN', 'Tong CBM CO HE SO', 'CBM CHUA HE SO']

        def render_summary_sheet(ws, summary_rows, title_text):
            last_col_letter = get_column_letter(total_data_cols)
            ws.merge_cells(f'A1:{last_col_letter}1')
            ws['A1'] = title_text
            ws['A1'].font = Font(name='Times New Roman', size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

            for col, header in enumerate(base_headers, start=1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(name='Times New Roman', size=11, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = fill_header_left
                cell.border = thin_border

            for idx, cfg in enumerate(account_configs, start=8):
                cell = ws.cell(row=3, column=idx, value=cfg['title'])
                cell.font = Font(name='Times New Roman', size=11, bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.fill = PatternFill(fill_type='solid', fgColor=cfg['color'])
                cell.border = thin_border

            ws.merge_cells('A4:G4')
            ws['A4'] = 'H? S?'
            ws['A4'].font = Font(name='Times New Roman', size=11, bold=True)
            ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A4'].fill = fill_header_left

            for idx, cfg in enumerate(account_configs, start=8):
                cell = ws.cell(row=4, column=idx, value=cfg['coef'])
                cell.font = Font(name='Times New Roman', size=11, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                cell.number_format = '0.0'

            row_count = max(15, len(summary_rows))
            for i in range(1, row_count + 1):
                row_no = 4 + i
                row_data = summary_rows[i - 1] if i - 1 < len(summary_rows) else None

                ws.cell(row=row_no, column=1, value=i)
                ws.cell(row=row_no, column=2, value=row_data['position'] if row_data else '')
                ws.cell(row=row_no, column=3, value=row_data['employee_code'] if row_data else '')
                ws.cell(row=row_no, column=4, value=row_data['masl'] if row_data else '')
                ws.cell(row=row_no, column=5, value=row_data['full_name'] if row_data else '')
                ws.cell(row=row_no, column=6, value=row_data['total_converted'] if row_data else 0)
                ws.cell(row=row_no, column=7, value=row_data['total_raw'] if row_data else 0)

                for j, cfg in enumerate(account_configs, start=8):
                    ws.cell(row=row_no, column=j, value=(row_data[cfg['title']] if row_data else 0))

                for c in range(1, total_data_cols + 1):
                    cell = ws.cell(row=row_no, column=c)
                    cell.font = Font(name='Times New Roman', size=11)
                    cell.border = thin_border
                    if c == 5:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    if c in (6, 7) or c >= 8:
                        cell.number_format = '#,##0.00'

                if 9 <= row_no <= 15:
                    for c in (2, 3, 4, 5):
                        ws.cell(row=row_no, column=c).fill = fill_yellow

            for r in range(3, row_count + 5):
                for c in range(1, total_data_cols + 1):
                    ws.cell(row=r, column=c).border = thin_border

            ws.column_dimensions['E'].width = 16.45
            ws.column_dimensions['F'].width = 23.73
            ws.column_dimensions['L'].width = 17.36
            ws.column_dimensions['N'].width = 16.27

            for c in range(1, total_data_cols + 1):
                col = get_column_letter(c)
                if ws.column_dimensions[col].width is None:
                    ws.column_dimensions[col].width = 12.5
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 11
            ws.column_dimensions['C'].width = 11
            ws.column_dimensions['D'].width = 13

            for r in range(1, row_count + 5):
                if r == 1:
                    ws.row_dimensions[r].height = 15
                elif r == 2:
                    ws.row_dimensions[r].height = 15.5
                elif r in (3, 4):
                    ws.row_dimensions[r].height = 14.5
                else:
                    ws.row_dimensions[r].height = 15.5

            ws.freeze_panes = 'A5'

        ws_khoan = wb.create_sheet('SAN_LUONG_KHOAN', 0)
        render_summary_sheet(ws_khoan, summary_rows_khoan, 'TONG HOP SAN LUONG THANG - KHOAN')

        ws_an_chung = wb.create_sheet('SAN_LUONG_AN_CHUNG', 1)
        render_summary_sheet(ws_an_chung, summary_rows_an_chung, 'TONG HOP SAN LUONG THANG - AN CHUNG')

        df_summary_template_khoan.to_excel(writer, index=False, sheet_name='TongHopKhoanRaw')
        df_summary_template_an_chung.to_excel(writer, index=False, sheet_name='TongHopAnChungRaw')
        df_detail.to_excel(writer, index=False, sheet_name='ChiTiet')
        if not df_anchung.empty:
            df_anchung.to_excel(writer, index=False, sheet_name='An_Chung_ChiTiet')

    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"report_{datetime.now().strftime('%Y%m%d')}.xlsx")

@app.route('/report/export-anchung')
@login_required
@view_required
def export_anchung():
    if not current_user.can_export:
        flash('Bạn không có quyền xuất báo cáo.', 'danger')
        return redirect(url_for('report'))
        
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    
    query = LaborProductivity.query
    if from_date: query = query.filter(LaborProductivity.work_date >= from_date)
    if to_date: query = query.filter(LaborProductivity.work_date <= to_date)
    
    records = query.order_by(LaborProductivity.work_date.desc()).all()
    
    an_chung_data = []
    an_chung_emps = Employee.query.filter_by(employee_type='An_chung').all()
    
    def normalize_key(s):
        if not s: return None
        return unicodedata.normalize('NFC', str(s)).strip().lower()

    ac_map = {}
    summary_map = {} # Map tổng hợp cho Excel

    for emp in an_chung_emps:
        if emp.masl: ac_map[normalize_key(emp.masl)] = emp
        if emp.employee_code: ac_map[normalize_key(emp.employee_code)] = emp
        if emp.full_name: ac_map[normalize_key(emp.full_name)] = emp
        
        # Khởi tạo dòng cho Excel
        summary_map[emp.employee_code] = {
            'Mã NV': emp.employee_code,
            'Mã SL': emp.masl,
            'Họ và tên': emp.full_name,
            'Vị trí': emp.position,
            'Số lượt tham gia': 0,
            'Số CBM chưa quy đổi': 0.0,
            'Số cbm đã quy đổi': 0.0
        }
        
    for r in records:
        workers = [
            r.tally_id, r.xenang_id, 
            r.congnhan1_id, r.congnhan2_id, r.congnhan3_id, 
            r.congnhan4_id, r.congnhan5_id, r.congnhan6_id
        ]
        seen_in_row = set()
        for w_str in workers:
            if not w_str: continue
            w_key = normalize_key(w_str)
            if w_key in ac_map:
                emp = ac_map[w_key]
                if emp.id in seen_in_row: continue
                seen_in_row.add(emp.id)
                
                an_chung_data.append({
                    'Ngày': r.work_date,
                    'Mã NV': emp.employee_code,
                    'Mã SL': emp.masl,
                    'Họ và tên': emp.full_name,
                    'Vị trí': emp.position,
                    'Task': r.task_id,
                    'Số CBM chưa quy đổi': r.productivity_value,
                    'Chỉ số quy đổi': r.conversion_index,
                    'Số cbm đã quy đổi': r.quantity
                })
                
                # Cộng dồn vào summary_map
                code = emp.employee_code
                if code in summary_map:
                    summary_map[code]['Số lượt tham gia'] += 1
                    summary_map[code]['Số CBM chưa quy đổi'] += (r.productivity_value or 0.0)
                    summary_map[code]['Số cbm đã quy đổi'] += (r.quantity or 0.0)

    df_anchung = pd.DataFrame(an_chung_data)
    
    # Tạo DataFrame tổng hợp từ summary_map (đầy đủ nhân viên)
    summary_list = list(summary_map.values())
    summary_list.sort(key=lambda x: x['Họ và tên'])
    df_summary = pd.DataFrame(summary_list)
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_summary.to_excel(writer, index=False, sheet_name='Tong_Hop')
        if not df_anchung.empty:
            df_anchung.to_excel(writer, index=False, sheet_name='Chi_Tiet')
        else:
            pd.DataFrame(['Không có dữ liệu chi tiết']).to_excel(writer, index=False, sheet_name='Chi_Tiet')

    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'AnChung_{datetime.now().strftime("%Y%m%d")}.xlsx')

@app.route('/import-data/template')
@login_required
@update_required
def download_template():
    # Định nghĩa các cột theo yêu cầu
    columns = [
        'Date', 'số cont/xe', 'cbm', 'tally', 'xe nang', 
        'cong nhan_1', 'cong nhan_2', 'cong nhan_3', 'cong nhan_4', 'cong nhan_5', 'cong nhan_6', 
        'task', 'account', 'khách hàng'
    ]
    
    # Lấy dữ liệu cho Combobox (Dropdown)
    # Lấy danh sách khách hàng
    customers = [c.customer_name for c in Customer.query.with_entities(Customer.customer_name).distinct().all() if c.customer_name]
    # Lấy danh sách account (tên)
    accounts = [a.account_name for a in CustomerAccount.query.with_entities(CustomerAccount.account_name).distinct().all() if a.account_name]
    # Lấy danh sách task (tên)
    tasks = [t.task_name for t in AccountTask.query.with_entities(AccountTask.task_name).distinct().all() if t.task_name]

    # Tạo DataFrame rỗng với các cột này
    df = pd.DataFrame(columns=columns)
    
    # Ghi vào bộ nhớ đệm (buffer) thay vì lưu ra ổ cứng
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Template')
        
        # Thêm Data Validation (Combobox)
        workbook = writer.book
        ws = writer.sheets['Template']
        
        # Tạo sheet ẩn chứa dữ liệu danh sách
        ws_data = workbook.create_sheet('DataList')
        ws_data.sheet_state = 'hidden'
        
        # Ghi dữ liệu vào sheet ẩn (Cột A: Khách hàng, B: Account, C: Task)
        for i, val in enumerate(customers, 1): ws_data.cell(row=i, column=1, value=val)
        for i, val in enumerate(accounts, 1): ws_data.cell(row=i, column=2, value=val)
        for i, val in enumerate(tasks, 1): ws_data.cell(row=i, column=3, value=val)

        # Hàm helper để thêm validation
        def add_validation(col_letter, data_len, col_idx_in_data):
            if data_len > 0:
                # Tạo tham chiếu đến sheet DataList (VD: 'DataList'!$A$1:$A$10)
                col_char = chr(64 + col_idx_in_data) # 1->A, 2->B, 3->C
                formula = f"'DataList'!${col_char}$1:${col_char}${data_len}"
                dv = DataValidation(type="list", formula1=formula, allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}1000") # Áp dụng cho 1000 dòng

        # Áp dụng validation cho các cột tương ứng
        # Task (L), Account (M), Khách hàng (N)
        add_validation('L', len(tasks), 3)      # Task lấy từ cột C (3) của DataList
        add_validation('M', len(accounts), 2)   # Account lấy từ cột B (2) của DataList
        add_validation('N', len(customers), 1)  # Khách hàng lấy từ cột A (1) của DataList

    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name='import_template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/import-data-view')
@login_required
@update_required
def import_data_view():
    return redirect(url_for('import_data'))

@app.cli.command("seed-db")
def seed_db():
    """Thêm dữ liệu chức vụ ban đầu vào database."""
    chuc_vu_list = ['Kiểm đếm', 'Tài xế xe nâng thấp', 'Tài xế xe nâng cao', 'Bốc xếp']
    
    print("Bắt đầu thêm dữ liệu cho bảng ChucVu...")
    for cv_name in chuc_vu_list:
        # Kiểm tra xem chức vụ đã tồn tại chưa
        exists = ChucVu.query.filter_by(ten_chuc_vu=cv_name).first()
        if not exists:
            new_cv = ChucVu(ten_chuc_vu=cv_name)
            db.session.add(new_cv)
            print(f"  -> Đã thêm: {cv_name}")
        else:
            print(f"  -> Bỏ qua (đã tồn tại): {cv_name}")
            
    db.session.commit()
    print("Hoàn tất!")

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']

        if not check_password_hash(current_user.password_hash, current_password):
            flash('Mật khẩu hiện tại không đúng.', 'danger')
        elif new_password != confirm_password:
            flash('Mật khẩu mới nhập lại không khớp.', 'danger')
        else:
            current_user.password_hash = generate_password_hash(new_password)
            db.session.commit()
            flash('Đổi mật khẩu thành công!', 'success')
            return redirect(url_for('index'))
    
    return render_template('change_password.html')

def create_default_admin():
    """Tạo user admin mặc định hoặc reset mật khẩu nếu cần."""
    try:
        # Kiểm tra xem user admin đã tồn tại chưa
        admin = User.query.filter_by(username='admin').first()
        
        if admin:
            print("User admin đã tồn tại. Đang reset mật khẩu về '123'...")
            admin.password_hash = generate_password_hash('123')
            db.session.commit()
            print("--> Đã reset mật khẩu thành công! Tài khoản: admin | Mật khẩu: 123")
        else:
            print("Chưa có user admin. Đang tạo tài khoản ADMIN mặc định...")
            hashed_pw = generate_password_hash('123')
            admin_user = User(
                username='admin',
                password_hash=hashed_pw,
                full_name='Administrator',
                role='ADMIN',
                is_active=True
            )
            db.session.add(admin_user)
            db.session.commit()
            print("--> Đã tạo thành công! Tài khoản: admin | Mật khẩu: 123")
    except Exception as e:
        print(f"Lỗi khi tạo admin mặc định (có thể do chưa tạo bảng): {e}")

if __name__ == '__main__':
    # Tạo admin mặc định trong context của ứng dụng
    # with app.app_context():
    #     create_default_admin()

    # Chạy ứng dụng
    app.run(host='0.0.0.0', port=5000)
